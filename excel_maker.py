import openpyxl as opxl
# TODO: Learn Pandas and Dataframes to figure out how to convert the input into usable data.


def create_input_workbook(num_breaks):
    book = opxl.Workbook()
    break_namer = 1
    for i in range(1, num_breaks + 1):
        book.create_sheet('Break ' + str(break_namer))
        break_namer += 1

    # Remove the default sheet created by openpyxl.
    del book['Sheet']

    # Acquire a sheet by its name.
    # sheet_break1 = book['Break 1']
    # sheet_sample = book['Sample']
    #
    # # Merging first 3 columns of default sheet.
    # r1 = 1
    # r2 = 3
    # c1 = 1
    # c2 = 3
    # sheet_sample.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    #
    # # Writing to sheet.
    # sheet_break1.cell(row=1, column=1).value = 'Name'
    # sheet_break1.cell(row=1, column=2).value = 'Acres'
    # sheet_break1.cell(row=1, column=3).value = 'NRI'
    #
    # # To set alignment of text inside cell and text wrapping.
    # sheet_break1.cell(row=1, column=1).alignment = opxl.styles.Alignment(horizontal='center', vertical='center',
    #                                                               wrap_text=True)
    # sheet_break1.cell(row=1, column=2).alignment = opxl.styles.Alignment(horizontal='center', vertical='center',
    #                                                               wrap_text=True)
    # sheet_break1.cell(row=1, column=3).alignment = opxl.styles.Alignment(horizontal='center', vertical='center',
    #                                                               wrap_text=True)
    #
    # # To make font bold or italic.
    # sheet_break1.cell(row=1, column=1).font = opxl.styles.Font(bold=True, underline='single')
    # sheet_break1.cell(row=1, column=2).font = opxl.styles.Font(bold=True, underline='single')
    # sheet_break1.cell(row=1, column=3).font = opxl.styles.Font(bold=True, underline='single')
    #
    book.save('Working Interest Input.xlsx')


create_input_workbook(3)

