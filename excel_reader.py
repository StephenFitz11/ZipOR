import openpyxl as opxl


def main():
    book = opxl.load_workbook('Sample.xlsx')

    # Get a sheet to read
    sheet = book['Break 1']

    # No. of rows written to a sheet.
    r = sheet.max_row

    # No. of columns written to a sheet.
    c = sheet.max_column

    # Reading each cell in Excel.
    for i in range(1, r+1):
        for j in range(1, c+1):
            print(sheet.cell(row=i, column=j).value)


if __name__ == '__main__':
    main()
