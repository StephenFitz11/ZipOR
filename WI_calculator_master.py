import gui_modules as gui
import pandas as pd
import openpyxl as opxl
import os

power = True

# GUI prompts user to select their desktop folder and changes os dir to that.
desktop = gui.ask_desktop()
os.chdir(str(desktop))

# GUI asks user to specify how many breaks there will be and creates a workbook with that many sheets.
break_count = gui.ask_depths()
book = opxl.Workbook()
break_namer = 1
for i in range(1, break_count + 1):
    book.create_sheet('Break ' + str(break_namer))
    break_namer += 1
del book['Sheet']  # Deletes the default sheet created by openpyxl

# Formats each sheet to list 'Name, Acres, NRI', on columns A, B, C, respectively.
for tab in range(1, break_count + 1):
    x = book['Break ' + str(tab)]
    x.cell(row=1, column=1).value = 'Name'
    x.cell(row=1, column=2).value = 'Acres'
    x.cell(row=1, column=3).value = 'NRI'

# Saves the workbook to the path.
path = r'data.xlsx'
book.save(path)

# GUI provides instructions to user.
gui.gui_instructions()

# Calculates the output of each tab in the file.
for sheet in range(1, break_count + 1):
    # Reads obj and coverts to Dataframe.
    read_data = pd.read_excel(path, sheet_name='Break ' + str(sheet))
    df = pd.DataFrame(read_data)

    # Calculates the royalty acres and adds that to the df as 'Royac' column.
    df['Royac'] = df['Acres'] * df['NRI']

    # Creates a list of all companies on the sheet. (Occurrences of each company is one).
    all_companies = []
    for company in df['Name']:
        all_companies.append(company)
    all_companies = list(dict.fromkeys(all_companies))

    # Creates a new Dataframe (df2) from the list to store output.
    df2 = pd.DataFrame(all_companies, columns=['Name'])

    # Creates lists of the calculated weighted NRIs and total acres for each company in df2.
    w_nris = []
    total_acres = []
    for company in df2['Name']:
        batch = df.loc[df['Name'] == company]
        weighted_nri = batch['Royac'].sum() / batch['Acres'].sum()
        w_nris.append(weighted_nri)
        total_acres.append(batch['Acres'].sum())

    # Appends the weighted NRIs and total acres to df2.
    df2['Net Acres'] = total_acres
    df2['Weighted NRIS'] = w_nris

    # Creates a new sheet in path and writes df2 to it.
    workbook = opxl.load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = workbook
    df2.to_excel(writer, sheet_name='Break ' + str(sheet) + ' OUTPUT')
    writer.save()
    writer.close()

gui.gui_complete()








